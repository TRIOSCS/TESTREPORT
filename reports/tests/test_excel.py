import pytest
import tempfile
import os
from django.test import TestCase
from reports.services.excel import ExcelWriter


class TestExcelWriter(TestCase):
    """Test Excel and CSV writing functionality"""
    
    def setUp(self):
        self.writer = ExcelWriter()
    
    def test_create_dataframe_empty(self):
        """Test creating DataFrame with empty data"""
        df = self.writer._create_dataframe([])
        
        self.assertEqual(len(df), 0)
        self.assertEqual(list(df.columns), self.writer.required_columns)
    
    def test_create_dataframe_with_data(self):
        """Test creating DataFrame with drive data"""
        drives = [
            {
                'VPD Serial': 'WD12345678901234567890',
                'Model Number': 'WD Blue 1TB',
                'Health Score': 95,
                'File Name': 'test.txt',
                'Vendor': 'Western Digital'
            },
            {
                'VPD Serial': 'ST98765432109876543210',
                'Model Number': 'ST1000DM010',
                'Health Score': 88,
                'File Name': 'test2.txt',
                'Vendor': 'Seagate'
            }
        ]
        
        df = self.writer._create_dataframe(drives)
        
        self.assertEqual(len(df), 2)
        self.assertEqual(df.iloc[0]['VPD Serial'], 'WD12345678901234567890')
        self.assertEqual(df.iloc[1]['VPD Serial'], 'ST98765432109876543210')
        self.assertEqual(df.iloc[0]['Vendor'], 'Western Digital')
        self.assertEqual(df.iloc[1]['Vendor'], 'Seagate')
    
    def test_write_excel_file(self):
        """Test writing Excel file"""
        drives = [
            {
                'VPD Serial': 'WD12345678901234567890',
                'Model Number': 'WD Blue 1TB',
                'Health Score': 95,
                'File Name': 'test.txt',
                'Vendor': 'Western Digital',
                'Label Serial': 'WD12345',
                'Vendor Information': 'Western Digital Corporation',
                'Connection / Interface Type': 'SATA',
                'Allocated Sections': 0,
                'Grown Defects': 0
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            result_path = self.writer.write_excel(drives, output_path)
            self.assertEqual(result_path, output_path)
            self.assertTrue(os.path.exists(output_path))
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_write_csv_file(self):
        """Test writing CSV file"""
        drives = [
            {
                'VPD Serial': 'WD12345678901234567890',
                'Model Number': 'WD Blue 1TB',
                'Health Score': 95,
                'File Name': 'test.txt',
                'Vendor': 'Western Digital',
                'Label Serial': 'WD12345',
                'Vendor Information': 'Western Digital Corporation',
                'Connection / Interface Type': 'SATA',
                'Allocated Sections': 0,
                'Grown Defects': 0
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            output_path = f.name
        
        try:
            result_path = self.writer.write_csv(drives, output_path)
            self.assertEqual(result_path, output_path)
            self.assertTrue(os.path.exists(output_path))
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_write_excel_with_errors(self):
        """Test writing Excel file with errors sheet"""
        drives = [
            {
                'VPD Serial': 'WD12345678901234567890',
                'Model Number': 'WD Blue 1TB',
                'Health Score': 95,
                'File Name': 'test.txt',
                'Vendor': 'Western Digital',
                'Label Serial': 'WD12345',
                'Vendor Information': 'Western Digital Corporation',
                'Connection / Interface Type': 'SATA',
                'Allocated Sections': 0,
                'Grown Defects': 0
            }
        ]
        
        errors = [
            {
                'file_name': 'bad_file.txt',
                'error_message': 'Could not parse file',
                'encodings_tried': ['utf-8', 'iso-8859-1']
            }
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            result_path = self.writer.write_excel(drives, output_path, errors)
            self.assertEqual(result_path, output_path)
            self.assertTrue(os.path.exists(output_path))
            
            # Check that file has multiple sheets
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            self.assertIn('Drive Summary', wb.sheetnames)
            self.assertIn('Errors', wb.sheetnames)
            
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
