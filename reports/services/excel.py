import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles Excel and CSV file generation with formatting"""
    
    def __init__(self):
        # Columns must appear in this exact order as specified
        self.required_columns = [
            "Label Serial",
            "VPD Serial",
            "Model Number", 
            "Vendor Information",
            "Vendor",
            "File Name",
            "Health Score",
            "Allocated Sections",
            "Grown Defects"
        ]
    
    def write_excel(self, drives: List[Dict[str, Any]], output_path: str, errors: List[Dict[str, Any]] = None) -> str:
        """
        Write drives data to Excel file with formatting.
        Returns the path to the created file.
        """
        try:
            # Create DataFrame
            df = self._create_dataframe(drives)
            
            # Create Excel file with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Drive Summary"
            
            # Write data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            self._apply_formatting(ws, df)
            
            # Add errors sheet if there are errors
            if errors:
                self._add_errors_sheet(wb, errors)
            
            # Save file
            wb.save(output_path)
            logger.info(f"Excel file saved to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise
    
    def write_csv(self, drives: List[Dict[str, Any]], output_path: str) -> str:
        """
        Write drives data to CSV file.
        Returns the path to the created file.
        """
        try:
            df = self._create_dataframe(drives)
            df.to_csv(output_path, index=False)
            logger.info(f"CSV file saved to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error writing CSV file: {e}")
            raise
    
    def _create_dataframe(self, drives: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create pandas DataFrame from drives data"""
        if not drives:
            # Create empty DataFrame with required columns
            return pd.DataFrame(columns=self.required_columns)
        
        # Ensure all drives have all required columns
        normalized_drives = []
        for drive in drives:
            normalized_drive = {}
            for col in self.required_columns:
                normalized_drive[col] = drive.get(col, '')
            normalized_drives.append(normalized_drive)
        
        df = pd.DataFrame(normalized_drives)
        
        # Ensure columns are in the correct order
        df = df[self.required_columns]
        
        return df
    
    def _apply_formatting(self, ws, df: pd.DataFrame):
        """Apply Excel formatting to the worksheet"""
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply conditional formatting for Health Score
        health_col_idx = self.required_columns.index("Health Score") + 1  # +1 for 1-based indexing
        health_col_letter = ws.cell(row=1, column=health_col_idx).column_letter
        
        # Define colors for conditional formatting
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light Yellow  
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light Red
        
        # Apply conditional formatting to Health Score column
        for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
            cell = ws.cell(row=row, column=health_col_idx)
            try:
                health_score = int(cell.value) if cell.value else 0
                
                if health_score > 95:
                    cell.fill = green_fill  # Green: >95%
                elif health_score >= 90:
                    cell.fill = yellow_fill  # Yellow: 90–95%
                elif health_score < 90:
                    cell.fill = red_fill  # Red: <90%
            except (ValueError, TypeError):
                # Skip cells that can't be converted to int
                pass
        
        # Bold header row
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    
    def _add_errors_sheet(self, wb: Workbook, errors: List[Dict[str, Any]]):
        """Add errors sheet to workbook"""
        ws_errors = wb.create_sheet("Errors")
        
        # Headers
        ws_errors.append(["File Name", "Error Details", "Encodings Attempted"])
        
        # Data
        for error in errors:
            ws_errors.append([
                error.get("file_name", ""),
                error.get("error_message", ""),
                ", ".join(error.get("encodings_tried", []))
            ])
        
        # Format errors sheet
        ws_errors.freeze_panes = "A2"
        
        # Auto-size columns
        for column in ws_errors.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws_errors.column_dimensions[column_letter].width = adjusted_width
        
        # Bold header row
        header_font = Font(bold=True)
        for cell in ws_errors[1]:
            cell.font = header_font
